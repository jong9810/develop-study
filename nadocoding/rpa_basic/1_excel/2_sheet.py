from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()  # 새로운 시트 생성
ws.title = "MySheet"  # 시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NewSheet", 2)  # 2번째 인덱스에 시트 생성(0부터 시작)

# 시트에 접근할 때
# 1) ws1.xxx
# 2) wb["YourSheet"] : Dict 형태로 시트에 접근 가능
new_ws = wb["NewSheet"]

# 모든 시트 이름 출력
print(wb.sheetnames)

# 시트 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample.xlsx")
wb.close()
